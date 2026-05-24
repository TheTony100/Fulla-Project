from __future__ import annotations

from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(size=10)

FILL_OK = PatternFill("solid", fgColor="C6EFCE")
FILL_LOW_CONF = PatternFill("solid", fgColor="FFEB9C")
FILL_NEEDS_REVIEW = PatternFill("solid", fgColor="FFC7CE")

_STATUS_LABELS: dict[str, str] = {
    "needs_review": "Needs Review",
    "low_confidence": "Low Confidence",
    "missing_field": "Missing Field",
    "ok": "OK",
    "processed": "Processed",
    "failed": "Failed",
    "unprocessed": "Unprocessed",
    "queued": "Queued",
    "merge_conflict": "Merge Conflict",
    "extraction_failed": "Extraction Failed",
    "missing_table_evidence": "Missing Table Evidence",
}


def format_status_label(value: object) -> str:
    """Convert snake_case status/issue codes to investor-facing labels."""
    if value is None:
        return ""
    text = str(value).strip()
    if not text or text.lower() in ("nan", "none"):
        return ""
    key = text.lower().replace(" ", "_")
    if key in _STATUS_LABELS:
        return _STATUS_LABELS[key]
    if "_" in text:
        return text.replace("_", " ").title()
    return text


def format_dataframe_labels(df: pd.DataFrame) -> pd.DataFrame:
    """Apply display labels to Review Status, Issue Type, and document Status columns."""
    if df.empty:
        return df
    out = df.copy()
    for col in ("Review Status", "Issue Type", "Status"):
        if col in out.columns:
            out[col] = out[col].apply(format_status_label)
    return out


def format_header_row(ws: Worksheet, *, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(row=row + 1, column=1).coordinate


def autosize_columns(ws: Worksheet, *, max_width: int = 48) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), max_width)


def apply_body_style(ws: Worksheet, *, start_row: int = 2) -> None:
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def hide_sheet(ws: Worksheet) -> None:
    ws.sheet_state = "hidden"


def _column_index(ws: Worksheet, header_name: str, *, header_row: int = 1) -> int | None:
    for idx, cell in enumerate(ws[header_row], start=1):
        if cell.value == header_name:
            return idx
    return None


def apply_review_status_formatting(
    ws: Worksheet,
    *,
    header_row: int = 1,
    status_column: str = "Review Status",
    issue_column: str | None = None,
) -> None:
    """Green = ok, yellow = low confidence, red = needs review / missing field."""
    if ws.max_row <= header_row:
        return
    status_col = _column_index(ws, status_column, header_row=header_row)
    issue_col = _column_index(ws, issue_column, header_row=header_row) if issue_column else None
    if status_col is None and issue_col is None:
        return

    data_start = header_row + 1
    data_end = ws.max_row
    last_col = get_column_letter(ws.max_column)
    cell_range = f"A{data_start}:{last_col}{data_end}"

    if issue_col:
        issue_letter = get_column_letter(issue_col)
        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(
                formula=[
                    f'OR(${issue_letter}{data_start}="Needs Review",'
                    f'${issue_letter}{data_start}="Missing Field")'
                ],
                fill=FILL_NEEDS_REVIEW,
            ),
        )
        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(
                formula=[f'${issue_letter}{data_start}="Low Confidence"'],
                fill=FILL_LOW_CONF,
            ),
        )

    if status_col:
        status_letter = get_column_letter(status_col)
        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(
                formula=[f'OR(${status_letter}{data_start}="OK",${status_letter}{data_start}="ok")'],
                fill=FILL_OK,
            ),
        )
        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(
                formula=[f'ISNUMBER(SEARCH("Low",{status_letter}{data_start}))'],
                fill=FILL_LOW_CONF,
            ),
        )
        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(
                formula=[f'ISNUMBER(SEARCH("Needs Review",{status_letter}{data_start}))'],
                fill=FILL_NEEDS_REVIEW,
            ),
        )


def finalize_worksheet(
    ws: Worksheet,
    *,
    header_row: int = 1,
    extra_header_row: int | None = None,
    status_column: str | None = "Review Status",
    issue_column: str | None = None,
    hidden: bool = False,
) -> None:
    if ws.max_row < 1:
        return
    format_header_row(ws, row=header_row)
    if extra_header_row is not None:
        format_header_row(ws, row=extra_header_row)
    apply_body_style(ws, start_row=header_row + 1)
    autosize_columns(ws)
    if status_column or issue_column:
        apply_review_status_formatting(
            ws,
            header_row=header_row,
            status_column=status_column or "Review Status",
            issue_column=issue_column,
        )
    if hidden:
        hide_sheet(ws)
